import os
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from datetime import datetime

WORKSPACE = r"C:\Users\tiger\Desktop"
WORKER_FILE = "Workers Detail.xlsx"
DEFAULT_EXCEL = "Work Schedule 3-8.xlsx"

st.set_page_config(
    page_title="Lee's Work Dashboard",
    page_icon="🔧",
    layout="wide",
)

# ── global table CSS ───────────────────────────────────────────────────────────
st.markdown("""
<style>
table.dash { width:100%; border-collapse:collapse; font-size:14px; }
table.dash th {
    background:#1e3a5f; color:white;
    padding:8px 10px; text-align:left;
    border:1px solid #ccc;
}
table.dash td { padding:7px 10px; border:1px solid #ddd; }
table.dash tr:nth-child(even) td { background:#f5f7fa; }
table.dash tr:hover td { background:#e8f0fe; }
</style>
""", unsafe_allow_html=True)


# ── helpers ────────────────────────────────────────────────────────────────────

def html_table(df, cell_styles=None):
    """Render a DataFrame as an HTML table (no pyarrow needed).
    cell_styles: dict of {(row_index, col_name): "css string"}
    """
    cols = list(df.columns)
    rows_html = ""
    for i, (_, row) in enumerate(df.iterrows()):
        cells = ""
        for col in cols:
            val = row[col]
            style = ""
            if cell_styles and (i, col) in cell_styles:
                style = f' style="{cell_styles[(i, col)]}"'
            cells += f"<td{style}>{val}</td>"
        rows_html += f"<tr>{cells}</tr>"

    header = "".join(f"<th>{c}</th>" for c in cols)
    return (
        f'<div style="overflow-x:auto">'
        f'<table class="dash"><thead><tr>{header}</tr></thead>'
        f'<tbody>{rows_html}</tbody></table></div>'
    )


def load_workers():
    path = os.path.join(WORKSPACE, WORKER_FILE)
    if not os.path.exists(path):
        return None
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[5]:
            continue
        rows.append({
            "WP No":            row[1],
            "FIN No":           row[2],
            "Nationality":      row[3],
            "Company":          row[4],
            "Worker Name":      row[5],
            "DOB":              row[6],
            "WP Expiry":        row[7],
            "WP Days Left":     row[8],
            "WP Status":        row[9],
            "CSOC Expiry":      row[10],
            "CSOC Days Left":   row[11],
            "CSOC Status":      row[12],
            "Coretrade No":     row[13],
            "Coretrade Expiry": row[14],
            "Coretrade Days":   row[15],
            "Coretrade Status": row[16],
        })
    return pd.DataFrame(rows)


def load_schedule():
    path = os.path.join(WORKSPACE, DEFAULT_EXCEL)
    if not os.path.exists(path):
        return None
    try:
        return pd.read_excel(path, header=0)
    except Exception as e:
        st.error(f"Schedule read error: {e}")
        return None


def days_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def badge(val):
    d = days_int(val)
    if d is None:
        return "–"
    if d <= 0:
        return f"❌ {d}d"
    if d <= 60:
        return f"🔴 {d}d"
    if d <= 90:
        return f"⚠️ {d}d"
    return f"✅ {d}d"


def days_bg(val):
    d = days_int(val)
    if d is None:
        return ""
    if d <= 0:
        return "background:#ffb3b3"
    if d <= 60:
        return "background:#ffd6a5"
    if d <= 90:
        return "background:#fff3b0"
    return ""


# ── page header ────────────────────────────────────────────────────────────────

st.title("🔧 Lee's Work Dashboard")
st.caption(f"Last refreshed: {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
if st.button("🔄 Refresh"):
    st.rerun()

# ── load data ──────────────────────────────────────────────────────────────────

workers_df = load_workers()
schedule_df = load_schedule()

# ── top metrics ────────────────────────────────────────────────────────────────

col1, col2, col3, col4 = st.columns(4)

if workers_df is not None:
    total = len(workers_df)

    def _count(col):
        expired = urgent = 0
        for v in workers_df[col]:
            d = days_int(v)
            if d is None:
                continue
            if d <= 0:
                expired += 1
            elif d <= 60:
                urgent += 1
        return expired, urgent

    wp_exp,  wp_urg  = _count("WP Days Left")
    cs_exp,  cs_urg  = _count("CSOC Days Left")
    ct_exp,  ct_urg  = _count("Coretrade Days")

    col1.metric("👷 Total Workers", total)
    col2.metric("🪪 WP Alerts",       wp_exp + wp_urg,  delta=f"{wp_exp} expired",  delta_color="inverse")
    col3.metric("📋 CSOC Alerts",     cs_exp + cs_urg,  delta=f"{cs_exp} expired",  delta_color="inverse")
    col4.metric("🏗️ Coretrade Alerts", ct_exp + ct_urg, delta=f"{ct_exp} expired",  delta_color="inverse")
else:
    col1.warning(f"`{WORKER_FILE}` not found in workspace.")

st.divider()

# ── tabs ───────────────────────────────────────────────────────────────────────

tab1, tab2, tab3, tab4 = st.tabs([
    "🚨 Renewal Alerts",
    "👷 All Workers",
    "📅 Work Schedule",
    "📁 Files",
])


# ── TAB 1 – Renewal Alerts ────────────────────────────────────────────────────
with tab1:
    if workers_df is None:
        st.warning(f"`{WORKER_FILE}` not found.")
    else:
        alerts = []
        for _, row in workers_df.iterrows():
            name = row["Worker Name"]
            for item, days_col in [
                ("WP",        "WP Days Left"),
                ("CSOC",      "CSOC Days Left"),
                ("Coretrade", "Coretrade Days"),
            ]:
                d = days_int(row[days_col])
                if d is None or d > 90:
                    continue
                if d <= 0:
                    status = "Expired"
                elif d <= 60:
                    status = "Urgent"
                else:
                    status = "Renew Soon"
                alerts.append({
                    "Worker":    name,
                    "Type":      item,
                    "Days Left": d,
                    "Status":    status,
                })

        if not alerts:
            st.success("✅ All passes are valid and above 90 days.")
        else:
            alert_df = (
                pd.DataFrame(alerts)
                .sort_values("Days Left")
                .reset_index(drop=True)
            )

            STATUS_STYLE = {
                "Expired":    "background:#ffb3b3;color:#8b0000;font-weight:bold",
                "Urgent":     "background:#ffd6a5;color:#7a3900;font-weight:bold",
                "Renew Soon": "background:#fff3b0;color:#7a6000",
            }

            cell_styles = {
                (i, "Status"): STATUS_STYLE.get(row["Status"], "")
                for i, row in alert_df.iterrows()
            }

            st.markdown(html_table(alert_df, cell_styles), unsafe_allow_html=True)
            st.write("")

            c1, c2, c3 = st.columns(3)
            exp  = sum(1 for a in alerts if a["Status"] == "Expired")
            urg  = sum(1 for a in alerts if a["Status"] == "Urgent")
            soon = sum(1 for a in alerts if a["Status"] == "Renew Soon")
            c1.error(f"❌ Expired: {exp}")
            c2.warning(f"🔴 Urgent (≤60 days): {urg}")
            c3.info(f"⚠️ Renew Soon (≤90 days): {soon}")


# ── TAB 2 – All Workers ───────────────────────────────────────────────────────
with tab2:
    if workers_df is None:
        st.warning(f"`{WORKER_FILE}` not found.")
    else:
        search = st.text_input("🔍 Search worker name / FIN", "")
        df = workers_df.copy()
        if search:
            mask = (
                df["Worker Name"].astype(str).str.contains(search, case=False, na=False) |
                df["FIN No"].astype(str).str.contains(search, case=False, na=False)
            )
            df = df[mask].reset_index(drop=True)

        display = pd.DataFrame({
            "Worker Name":      df["Worker Name"],
            "FIN No":           df["FIN No"],
            "Nationality":      df["Nationality"],
            "Company":          df["Company"],
            "WP Expiry":        df["WP Expiry"],
            "WP":               df["WP Days Left"].apply(badge),
            "CSOC Expiry":      df["CSOC Expiry"],
            "CSOC":             df["CSOC Days Left"].apply(badge),
            "Coretrade Expiry": df["Coretrade Expiry"],
            "Coretrade":        df["Coretrade Days"].apply(badge),
        })

        # colour the badge cells
        cell_styles = {}
        for i, row in df.iterrows():
            for badge_col, days_col in [
                ("WP",        "WP Days Left"),
                ("CSOC",      "CSOC Days Left"),
                ("Coretrade", "Coretrade Days"),
            ]:
                style = days_bg(row[days_col])
                if style:
                    cell_styles[(i, badge_col)] = style

        st.markdown(html_table(display, cell_styles), unsafe_allow_html=True)
        st.caption(f"{len(display)} worker(s) shown")


# ── TAB 3 – Work Schedule ─────────────────────────────────────────────────────
with tab3:
    if schedule_df is None:
        st.warning(f"`{DEFAULT_EXCEL}` not found.")
    else:
        st.subheader("Work Schedule")

        fc1, fc2 = st.columns(2)

        leader_col  = schedule_df.columns[5] if len(schedule_df.columns) > 5 else None
        project_col = schedule_df.columns[1] if len(schedule_df.columns) > 1 else None

        sdf = schedule_df.copy().reset_index(drop=True)

        if leader_col:
            leaders = ["All"] + sorted(sdf[leader_col].dropna().unique().tolist())
            sel_leader = fc1.selectbox("Filter by Leader", leaders)
            if sel_leader != "All":
                sdf = sdf[sdf[leader_col] == sel_leader].reset_index(drop=True)

        if project_col:
            projects = ["All"] + sorted(sdf[project_col].dropna().unique().tolist())
            sel_project = fc2.selectbox("Filter by Project", projects)
            if sel_project != "All":
                sdf = sdf[sdf[project_col] == sel_project].reset_index(drop=True)

        sdf = sdf.fillna("").astype(str)
        st.markdown(html_table(sdf), unsafe_allow_html=True)
        st.caption(f"{len(sdf)} row(s) shown")


# ── TAB 4 – Files ─────────────────────────────────────────────────────────────
with tab4:
    st.subheader("Files in Workspace")
    st.caption(WORKSPACE)

    try:
        file_rows = []
        for fname in sorted(os.listdir(WORKSPACE)):
            full = os.path.join(WORKSPACE, fname)
            if not os.path.isfile(full):
                continue
            size  = os.path.getsize(full)
            mtime = datetime.fromtimestamp(os.path.getmtime(full))
            ext   = os.path.splitext(fname)[1].lower() or "(no ext)"
            file_rows.append({
                "File":       fname,
                "Type":       ext,
                "Size (KB)":  round(size / 1024, 1),
                "Modified":   mtime.strftime("%d %b %Y %H:%M"),
            })

        fdf = pd.DataFrame(file_rows)

        ext_filter = st.selectbox(
            "Filter by type",
            ["All"] + sorted(fdf["Type"].unique().tolist()),
        )
        if ext_filter != "All":
            fdf = fdf[fdf["Type"] == ext_filter].reset_index(drop=True)

        st.markdown(html_table(fdf), unsafe_allow_html=True)
        st.caption(f"{len(fdf)} file(s)")

    except Exception as e:
        st.error(f"Could not read workspace: {e}")
